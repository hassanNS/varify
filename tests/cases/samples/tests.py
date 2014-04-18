import json
import os
from openpyxl import Workbook
from restlib2.http import codes
from ..base import AuthenticatedBaseTestCase


class VariantUploadResourceTestCase(AuthenticatedBaseTestCase):
    fixtures = ['initial_variants.json']

    def setUp(self):
        super(VariantUploadResourceTestCase, self).setUp()

    def test_post(self):
        book = Workbook()
        sheet1 = book.get_active_sheet()
        sheet1.title = 'Variants List'
        fields = ['Chromosome', 'Start', 'Reference', 'Allele 1', 'Allele 2',
                  'dbSNP']

        # Create variants to cover all edge cases, including the case
        # where there is a variation at the same genomic position.
        # Also consider the case where the dbSNP id is incorrect.
        variants = [['1', '20000', '.', 'AC', 'AC', ''],
                    ['1', '20000', 'A', 'A', '.', ''],
                    ['3', '20002', 'GAT', '.', '.', 'rs9160301'],
                    ['1', '20003', '.', '.', 'TTTCTT', ''],
                    ['3', '20004', 'A', 'C', 'C', 'rs916000'],
                    ['1', '20007', 'GTCATTGGAACAGTC', '.',
                     'GTCATTGGAACAGTC', '']]

        # First write the fields to our file across 4 columns
        for i in range(0, 6):
            sheet1.cell(row=0, column=i).value = fields[i]

        # Write the variants to the excel sheet
        row = 1
        for v in variants:
            for col in range(0, 6):
                sheet1.cell(row=row, column=col).value = str(v[col])
            row += 1

        book.save('variantList.xlsx')

        with open('variantList.xlsx') as fp:
            response = self.client.post('/api/samples/uploadFile/',
                                        {'sampleName': 'NA999',
                                         'name': 'variants',
                                         'qqfile': fp})
            response_obj = json.loads(response.content)
            # An array of matching variants are returned. Make sure
            # that exactly 5 were found and 1 was added to the
            # list of unmatched variants
            self.assertEqual(response.status_code, codes.ok)
            self.assertEqual(len(response_obj['results']), 5)
            self.assertEqual(len(response_obj['noResults']), 1)
        os.remove('variantList.xlsx')
