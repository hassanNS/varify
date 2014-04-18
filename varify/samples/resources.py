import functools
import openpyxl
import logging
import traceback
import requests
import sys
from restlib2.http import codes
from django.core import management
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf.urls import patterns, url
from django.core.cache import cache
from django.http import HttpResponse, Http404
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse
from django.views.decorators.cache import never_cache
from django.conf import settings
from preserialize.serialize import serialize
from restlib2 import resources
from serrano.resources.base import ThrottledResource
from varify.variants.resources import VariantResource
from varify import api
from varify.assessments.models import Assessment
from .models import Sample, Result, ResultScore
from varify.variants.models import Variant
from varify.genome.models import Chromosome

log = logging.getLogger(__name__)


def sample_posthook(instance, data, request):
    uri = request.build_absolute_uri
    data['_links'] = {
        'self': {
            'rel': 'self',
            'href': uri(reverse('api:samples:sample', args=[instance.pk])),
        },
        'variants': {
            'rel': 'related',
            'href': uri(reverse('api:samples:variants', args=[instance.pk])),
        }
    }

    return data


class SampleResource(ThrottledResource):
    model = Sample

    template = api.templates.Sample

    def is_not_found(self, request, response, pk):
        return not self.model.objects.filter(pk=pk).exists()

    @api.cache_resource
    def get(self, request, pk):
        try:
            sample = self.model.objects.select_related(
                'batch', 'project').get(pk=pk)
        except self.model.DoesNotExist:
            raise Http404

        posthook = functools.partial(sample_posthook, request=request)
        return serialize(sample, posthook=posthook, **self.template)


class SamplesResource(ThrottledResource):
    model = Sample

    template = api.templates.Sample

    def get(self, request):
        samples = self.model.objects.all()
        posthook = functools.partial(sample_posthook, request=request)
        return serialize(samples, posthook=posthook, **self.template)


class NamedSampleResource(ThrottledResource):
    "Resource for looking up a sample by project, batch, and sample name"
    model = Sample

    template = api.templates.Sample

    # Bypass authorization check imposed by Serrano's AUTH_REQUIRED setting
    def __call__(self, *args, **kwargs):
        return resources.Resource.__call__(self, *args, **kwargs)

    def is_not_found(self, request, response, project, batch, sample):
        try:
            instance = self.model.objects.get(project__name=project,
                                              batch__name=batch, name=sample)
        except self.model.DoesNotExist:
            return True

        request.instance = instance
        return False

    def get(self, request, project, batch, sample):
        data = serialize(request.instance, **self.template)
        data['_links'] = {
            'self': {
                'rel': 'self',
                'href': reverse('api:samples:sample',
                                kwargs={'pk': request.instance.pk})
            },
            'variants': {
                'rel': 'related',
                'href': reverse('api:samples:variants',
                                kwargs={'pk': request.instance.pk}),
            }
        }
        return data


class SampleResultsResource(ThrottledResource):
    "Paginated view of results for a sample."
    model = Result

    template = api.templates.SampleResult

    def is_not_found(self, request, response, pk):
        return not Sample.objects.filter(pk=pk).exists()

    def get(self, request, pk):
        page = request.GET.get('page', 1)

        related = ['sample', 'variant', 'variant__chr', 'genotype']
        results = self.model.objects.select_related(*related)\
            .filter(sample__pk=pk)

        # Paginate the results
        paginator = Paginator(results, api.PAGE_SIZE)

        try:
            page = page = paginator.page(page)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)

        resp = {
            'result_count': paginator.count,
            'results': serialize(page.object_list, **self.template),
        }

        # Augment the links
        for obj in resp['results']:
            obj['_links'] = {
                'self': {
                    'rel': 'self',
                    'href': reverse('api:samples:variant',
                                    kwargs={'pk': obj['id']})
                },
                'sample': {
                    'rel': 'related',
                    'href': reverse('api:samples:sample',
                                    kwargs={'pk': obj['sample']['id']})
                },
                'variant': {
                    'rel': 'related',
                    'href': reverse('api:variants:variant',
                                    kwargs={'pk': obj['variant_id']}),
                }
            }
            obj.pop('variant_id')

        links = {}
        if page.number != 1:
            links['prev'] = {
                'rel': 'prev',
                'href': "{0}?page={1}".format(
                    reverse('api:samples:variants', kwargs={'pk': pk}),
                    str(page.number - 1))
            }
        if page.number < paginator.num_pages - 1:
            links['next'] = {
                'rel': 'next',
                'href': "{0}?page={1}".format(
                    reverse('api:samples:variants', kwargs={'pk': pk}),
                    str(page.number + 1))
            }
        if links:
            resp['_links'] = links
        return resp


class SampleResultResource(ThrottledResource):
    model = Result

    template = api.templates.SampleResultVariant

    def is_not_found(self, request, response, pk):
        return not self.model.objects.filter(pk=pk).exists()

    def _cache_data(self, request, pk, key):
        related = ['sample', 'variant', 'genotype', 'score']

        try:
            result = self.model.objects.select_related(*related).get(pk=pk)
        except self.model.DoesNotExist:
            raise Http404

        data = serialize(result, **self.template)

        data['_links'] = {
            'self': {
                'rel': 'self',
                'href': reverse('api:samples:variant',
                                kwargs={'pk': data['id']})
            },
            'sample': {
                'rel': 'related',
                'href': reverse('api:samples:sample',
                                kwargs={'pk': data['sample']['id']})
            },
            'variant': {
                'rel': 'related',
                'href': reverse('api:variants:variant',
                                kwargs={'pk': data['variant_id']}),
            }
        }

        # Integrate the Variant resource data
        data['variant'] = VariantResource.get(request, data['variant_id'])
        data.pop('variant_id')

        try:
            score = ResultScore.objects.get(result=result)
            data['score'] = {
                'score': score.score,
                'rank': score.rank,
            }
        except ResultScore.DoesNotExist:
            pass

        cache.set(key, data, timeout=api.CACHE_TIMEOUT)
        return data

    def get(self, request, pk):
        key = api.cache_key(self.model, pk)
        data = cache.get(key)

        if data is None:
            data = self._cache_data(request, pk, key)

        try:
            assessment = Assessment.objects.get(sample_result=pk,
                                                user=request.user.id)
            data['assessment'] = serialize(assessment,
                                           **api.templates.ResultAssessment)
        except Assessment.DoesNotExist:
            data['assessment'] = {}

        return data


class PhenotypeResource(ThrottledResource):
    def get(self, request, sample_id):
        recalculate = request.GET.get('recalculate_rankings')

        if recalculate == "true":
            try:
                management.call_command('samples', 'gene-ranks', sample_id,
                                        force=True)
            except Exception:
                log.exception("Error recalculating gene rankings")
                return HttpResponse("Error recalculating gene rankings",
                                    status=500)

        endpoint = getattr(settings, 'PHENOTYPE_ENDPOINT', None)

        if not endpoint:
            log.error("PHENOTYPE_ENDPOINT setting could not be found.")
            return HttpResponse(status=500)

        endpoint = endpoint.format(sample_id)

        try:
            response = requests.get(endpoint, cert=(settings.VARIFY_CERT,
                                    settings.VARIFY_KEY), verify=False)
        except requests.exceptions.SSLError:
            raise PermissionDenied
        except requests.exceptions.ConnectionError:
            return HttpResponse(status=500)
        except requests.exceptions.RequestException:
            raise Http404

        # If anything at all goes wrong in the sample lookup or json parsing
        # then just abandon all hope and return the content from the orignal
        # response.
        try:
            sample = Sample.objects.get(label=sample_id)
            data = response.json()
            data['phenotype_modified'] = sample.phenotype_modified
            return data
        except Exception:
            pass

        return response.content


class PedigreeResource(ThrottledResource):
    def get(self, request, year, month, day, name):
        endpoint = getattr(settings, 'PEDIGREE_ENDPOINT', None)

        if not endpoint:
            log.error('PEDIGREE_ENDPOINT setting could not be found.')
            return HttpResponse(status=500)

        endpoint = endpoint.format(year, month, day, name)

        try:
            pedigree_response = requests.get(
                endpoint, cert=(settings.VARIFY_CERT, settings.VARIFY_KEY),
                verify=False)
        except requests.exceptions.SSLError:
            raise PermissionDenied
        except requests.exceptions.ConnectionError:
            return HttpResponse(status=500)
        except requests.exceptions.RequestException:
            raise Http404

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = \
            'attachment; filename="{0}-{1}-{2}-{3}"'.format(
                year, month, day, name)

        response.write(pedigree_response.content)

        return response


class Upload_Res(ThrottledResource):

    supported_content_types = ('multipart/form-data',
                               'application/vnd.openxmlformats-officedocument'
                               + 'spreadsheetml.sheet', )

    def change(self, ref, a1, a2):
        """ Given ref allele1 and allele2, returns the type of change """
        if ref == '.':
            return 'INSERTION'
        elif a1 == '.' or a2 == '.':
            return 'DELETION'
        else:
            return 'NOCHANGE'

    # Consume the variants excel file and search the database for
    # Matching queries
    def post(self, request):

        # Testing with sample NA12891 and batch1
        sampleName = request.POST['sampleName']
        #batchName = request.POST['batchName'].lower()
        #batch = Batch.objects.get(name=batchName)
        sample = Sample.objects.filter(name=sampleName)
        sample = sample[0]

        # Try opening the excel file that was uploaded
        try:
            variantBook = request.FILES['qqfile']
            workbook = openpyxl.load_workbook(variantBook)
        except:
            return HttpResponse(status=codes.unprocessable_entity)

        sheet = workbook.get_active_sheet()

        # Find the row where our useful data starts
        startRow = 0
        for row in range(0, len(sheet.rows)):
            cellValue = sheet.cell(row=row, column=0).value
            if cellValue == 'Chromosome':
                    startRow = row
                    break

        # retrive all the fields from the sheet
        fields = [sheet.cell(row=startRow, column=col).value
                  for col in range(len(sheet.columns))]

        fields = [field.lower() for field in fields]

        matches = set()
        noMatches = []

        for row in range(startRow + 1, len(sheet.rows)):
            # Retrive rsid, ref, start and chr from our spreadsheet
            # So we can retrive objects from our database
            dbsnpValue = sheet.cell(row=row,
                                    column=fields.index('dbsnp')).value
            dbsnpValue = dbsnpValue if dbsnpValue else None
            refValue = sheet.cell(row=row,
                                  column=fields.index('reference')).value
            startValue = sheet.cell(row=row,
                                    column=fields.index('start')).value
            chrLabel = \
                sheet.cell(row=row,
                           column=fields.index('chromosome')).value
            print chrLabel

            # Insertion, deletion or none?
            change_type = None

            chrMatch = Chromosome.objects.get(value=chrLabel)

            isFound = False
            # Fetch the variant w/ matching dbSNP id and the
            # Results which have the variant
            if dbsnpValue:
                try:
                    variant = Variant.objects.get(rsid=dbsnpValue)
                    result = Result.objects.filter(sample_id=sample.id,
                                                   variant_id=variant.id)
                    if len(result) > 0:
                        matches.update(result)
                        isFound = True
                except:
                    traceback.print_exc(file=sys.stdout)

            # In the case this form has a genotype field instead of allele1,2
            #elif 'genotype' in fields:
            #    # A form w/ genotype field should also have an alt field
            #    gtVcfValue = sheet.cell_value(row, fields.index('genotype'))
            #    altValue = sheet.cell_value(row, fields.index('alternate'))

            #    # Construct the genotype value based on our ref and alt
            #    #gtValue = ""
            #    #if gtVcfValue in ('1/1', '1/2'):
            #    #    gtValue = altValue + '/' + altValue
            #    #elif gtVcfValue == '0/1':
            #    #    gtValue = refValue + '/' + altValue
            #    #elif gtVcfValue == '0/0':
            #    #    gtValue = refValue + '/' + altValue

            #    genoMatch = Genotype.objects.get(value=gtVcfValue)
            #    variantMatch = Variant.objects.get(pos=startValue,
            #                                       ref=refValue,
            #                                       alt=altValue,
            #                                       chr_id=chrMatch.id)

            #    result = Result.objects.filter(genotype=genoMatch.id,
            #                                   variant_id=variantMatch.id)
            #    matches.update(result)
            if ('allele 1' in fields and isFound is False):
                # Retrive the two alleles and construct the genotype
                allele1 = sheet.cell(row=row,
                                     column=fields.index('allele 1')).value
                allele2 = sheet.cell(row=row,
                                     column=fields.index('allele 2')).value

                if allele1 == '.' or allele2 == '.' or refValue == '.':
                    # Regardless of an insertion or a deletion, we see that
                    # the start position decrements by 1. We will then query
                    # using the position only
                    startValue -= 1
                    # Determine the type of change
                    change_type = self.change(refValue, allele1, allele2)

                else:
                    # Now using the two alleles, construct the alt
                    altValue = ""
                    if allele1 == refValue:
                        altValue += allele2
                    else:
                        altValue += ''.join([allele1, ",", allele2]) \
                            if allele1 != allele2 else allele1

                # Retrive variant needed for our result query
                try:
                    correctVariant = ''

                    if change_type:
                        variantMatch = \
                            Variant.objects.filter(pos=startValue,
                                                   chr_id=chrMatch.id)
                        correctVariant = variantMatch[0]

                        if len(variantMatch) > 1:
                            for match in variantMatch:
                                if change_type == 'INSERTION' and \
                                        len(match.ref) == len(refValue) + 1:
                                    correctVariant = match
                                    break
                                elif change_type == 'DELETION' and \
                                        len(match.ref) == 1:
                                    correctVariant = match
                                    break

                    else:
                        variantMatch = Variant.objects.get(pos=startValue,
                                                           ref=refValue,
                                                           alt=altValue,
                                                           chr_id=chrMatch.id)
                        correctVariant = variantMatch

                    result = \
                        Result.objects.filter(sample_id=sample.id,
                                              variant_id=correctVariant.id)
                    if len(result) > 0:
                        matches.update(result)
                        isFound = True
                except:
                    pass

            # No method worked, failed to find result
            if isFound is False:
                allele1 = sheet.cell(row=row,
                                     column=fields.index('allele 1')).value
                allele2 = sheet.cell(row=row,
                                     column=fields.index('allele 2')).value
                # No matches for these rows. Save them to display to the user
                unMatchedFields = ['Chromosome:{0}, Start:{1}, Reference:{2}',
                                   ' Allele1:{3}, Allele2:{4}, dbsnp: {5}\n']
                unMatchedVariant = ''.join(unMatchedFields).format(chrLabel,
                                                                   startValue,
                                                                   refValue,
                                                                   allele1,
                                                                   allele2,
                                                                   dbsnpValue)

                noMatches.append(str(unMatchedVariant))

        # Strings that will display information about the results
        results = []
        for result in matches:
            detailString = ['Variant ID: {0}, Position: {1}, dbSNP: {2}\n ']
            details = ''.join(detailString).format(result.variant.id,
                                                   result.variant.pos,
                                                   result.variant.rsid)
            results.append(details)

        return {
            'results': results,
            'noResults': noMatches
        }


upload_resource = never_cache(Upload_Res())
sample_resource = never_cache(SampleResource())
samples_resource = never_cache(SamplesResource())
named_sample_resource = never_cache(NamedSampleResource())
sample_results_resource = never_cache(SampleResultsResource())
sample_result_resource = never_cache(SampleResultResource())
phenotype_resource = never_cache(PhenotypeResource())
pedigree_resource = never_cache(PedigreeResource())

urlpatterns = patterns(
    '',
    url(r'^$', samples_resource, name='samples'),
    url(r'^uploadFile/$', upload_resource, name='upload'),
    url(r'^(?P<pk>\d+)/$', sample_resource, name='sample'),
    url(r'^(?P<project>.+)/(?P<batch>.+)/(?P<sample>.+)/$',
        named_sample_resource, name='named_sample'),
    url(r'^(?P<pk>\d+)/variants/$', sample_results_resource, name='variants'),
    url(r'^variants/(?P<pk>\d+)/$', sample_result_resource, name='variant'),
    url(r'^(?P<sample_id>.+)/phenotypes/$', phenotype_resource,
        name='phenotype'),
    url(r'^pedigrees/(?P<year>\d+)/(?P<month>\d+)/(?P<day>\d+)/(?P<name>.+)$',
        pedigree_resource, name='pedigree'),
)
